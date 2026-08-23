"""Spreadsheet export/import: OpenPyXL (Excel) and EZSheets (Google Sheets)."""

from __future__ import annotations

import logging
import sys
from pathlib import Path
from tkinter import Tk, filedialog
from typing import Any

import ezsheets
import openpyxl
import pymsgbox

from . import config
from .database import SQLiteBillingDatabase
from .utils import generate_timestamp


class SpreadsheetExport:
    """Excel generation plus Google Sheets upload/download flows."""

    def __init__(self, spreadsheets_folder: Path, data_folder: Path | None = None) -> None:
        self.spreadsheets_folder = Path(spreadsheets_folder)
        self.data_folder = Path(data_folder) if data_folder else config.DATA_FOLDER
        self.current_path = self.spreadsheets_folder
        self.current_page = 0
        self.pages: list[Any] = []
        self.database_path = config.billing_database_path(self.data_folder)

    def _open_database(self) -> SQLiteBillingDatabase:
        return SQLiteBillingDatabase(self.database_path)

    def spreadsheet_interface_loop(self, system_tui: Any) -> None:
        system_tui.current_page = 0
        while True:
            system_tui.prepare_pages(self.current_path)
            system_tui.render_tui(
                "[1] Next | [2] Previous | [F] Download Sheets | [C] Upload Sheets | "
                "[X] Create Spreadsheet | [B] Back | [ESC] Exit",
                self.current_path,
            )

            option = input("Choose an option and press Enter: ").strip().lower()
            if option == "1" and system_tui.current_page < len(system_tui.pages) - 1:
                system_tui.current_page += 1
            elif option == "2" and system_tui.current_page > 0:
                system_tui.current_page -= 1
            elif option in ["b", "s"]:
                system_tui.current_page = 0
                break
            elif option == "esc":
                sys.exit()
            elif option == "x":
                self.create_spreadsheet()
            elif option == "c":
                self.upload_spreadsheet_prompt()
            elif option == "f":
                self.download_spreadsheet_prompt()

    def upload_spreadsheet_prompt(self) -> None:
        logging.info("Starting file picker for Google Sheets upload.")
        try:
            root = Tk()
            root.withdraw()
            root.attributes("-topmost", True)

            chosen_path = filedialog.askopenfilename(
                title="Select an Excel Spreadsheet to Upload",
                filetypes=[("Excel Files", "*.xlsx *.xls")],
            )
            root.destroy()

            if chosen_path:
                file_path = Path(chosen_path)
                if file_path.exists() and file_path.suffix in [".xlsx", ".xls"]:
                    self.create_google_sheet(file_path)
                else:
                    logging.warning("The selected file is invalid.")
                    pymsgbox.alert("Invalid file or unsupported format.", "Error")
            else:
                logging.info("User canceled file selection.")
        except Exception as exc:
            logging.error("Upload processing error: %s", exc, exc_info=True)
            pymsgbox.alert(f"Upload processing error: {exc}", "Error")

    def download_spreadsheet_prompt(self) -> None:
        logging.info("Starting Google Sheets download flow.")
        archive_types = ["xlsx", "ods", "csv", "pdf", "tsv", "zip"]
        for index, available_extension in enumerate(archive_types):
            print(f" [{index}] {available_extension.upper()}")

        type_choice = input("Choose the download file type: ")
        if type_choice.isdigit() and 0 <= int(type_choice) < len(archive_types):
            selected_type = archive_types[int(type_choice)]
            url = pymsgbox.prompt("Paste the Google Sheets URL to download:")

            if url is None:
                logging.info("User canceled the URL download operation.")
            elif url.strip():
                try:
                    logging.info("URL provided for download: %s", url)
                    self.download_sheet(url, selected_type)
                except Exception as exc:
                    logging.error("Download flow error: %s", exc, exc_info=True)
                    pymsgbox.alert(f"Download error:\n{exc}", "Error")
            else:
                logging.warning("No URL was provided by the user.")
                pymsgbox.alert("Invalid or empty URL.", "Error")
        else:
            logging.warning("Invalid download file type entered in the terminal.")
            pymsgbox.alert("Invalid file type.", "Error")

    def format_worksheet(self, worksheet: Any) -> None:
        try:
            worksheet.cell(row=1, column=1).value = "ACCOUNT"
            worksheet.cell(row=1, column=2).value = "PRODUCT ITEM"
            worksheet.cell(row=1, column=3).value = "PRICE"
            worksheet.cell(row=1, column=4).value = "REGISTRATION DATE / TIME"
            worksheet.cell(row=1, column=5).value = "CUSTOMER NAME"

            for column in range(7, 52, 4):
                cell = worksheet.cell(row=2, column=column)
                cell.value = f"--{worksheet.title.upper()} PANEL--"
            for column_id in range(9, 52, 4):
                cell = worksheet.cell(row=1, column=column_id)
                cell.value = "--POWERED BY LF SYSTEM--"
        except Exception as exc:
            logging.error("Worksheet formatting error: %s", exc)

    def create_spreadsheet(self) -> None:
        try:
            workbook = openpyxl.Workbook()
            profit_sheet = workbook.active
            profit_sheet.title = "Profit"
            workbook.create_sheet("Expenses")

            for worksheet in workbook.worksheets:
                self.format_worksheet(worksheet)

            with self._open_database() as database:
                all_accounts = database.get("accounts", {})
                if not all_accounts:
                    logging.warning("No accounts found for Excel export.")
                    pymsgbox.alert("No local database records were found to export.", "Warning")
                    return

                global_purchase_list = []
                for account_name, details in all_accounts.items():
                    history = details.get("purchase_history", [])
                    billing_items = details.get("billing_items", {})
                    customer_name = billing_items.get("customer_name", "N/A")

                    for purchase in history:
                        global_purchase_list.append(
                            {
                                "account": account_name,
                                "customer": customer_name,
                                "item": purchase["item"],
                                "price": purchase.get("price"),
                                "date": purchase.get("date", "N/A"),
                            }
                        )

                if not global_purchase_list:
                    logging.warning("No purchase history found for export.")
                    pymsgbox.alert("No purchase history has been registered.", "Warning")
                else:
                    ordered_purchases = sorted(
                        global_purchase_list,
                        key=lambda purchase: purchase["date"],
                        reverse=True,
                    )

                    for index, purchase in enumerate(ordered_purchases):
                        current_row = index + 3
                        profit_sheet.cell(row=current_row, column=1).value = purchase["account"]
                        profit_sheet.cell(row=current_row, column=2).value = purchase["item"]
                        profit_sheet.cell(row=current_row, column=4).value = purchase["date"]
                        profit_sheet.cell(row=current_row, column=5).value = purchase["customer"]

                        try:
                            price_cell = profit_sheet.cell(row=current_row, column=3)
                            price_cell.value = float(purchase["price"])
                            price_cell.number_format = '"$"#,##0.00'
                        except (ValueError, TypeError):
                            profit_sheet.cell(row=current_row, column=3).value = purchase["price"]

                    file_name = "LF_System_Export.xlsx"
                    save_path = self.spreadsheets_folder / file_name
                    workbook.save(save_path)
                    pymsgbox.alert(
                        "Individual audit spreadsheet generated successfully."
                        f"\nSaved as: {file_name}",
                        "Excel",
                    )
        except Exception as exc:
            logging.error("OpenPyXL error: %s", exc, exc_info=True)

    def download_sheet(self, url: str, archive_type: str) -> None:
        try:
            if "docs.google.com/spreadsheets" in url:
                url = url.split("/d/")[1].split("/edit")[0]

            sheets_workbook = ezsheets.Spreadsheet(url)
            file_name = f"Complete_Spreadsheet_{generate_timestamp()}.{archive_type}"
            save_path = self.spreadsheets_folder / file_name

            match archive_type.lower():
                case "xlsx":
                    sheets_workbook.downloadAsExcel(str(save_path))
                case "ods":
                    sheets_workbook.downloadAsODS(str(save_path))
                case "csv":
                    sheets_workbook.downloadAsCSV(str(save_path))
                case "tsv":
                    sheets_workbook.downloadAsTSV(str(save_path))
                case "zip":
                    sheets_workbook.downloadAsHTML(str(save_path))
                case "pdf":
                    sheets_workbook.downloadAsPDF(str(save_path))
                case _:
                    pymsgbox.alert("Invalid format.", "Error")
                    return

            pymsgbox.alert(
                f"File downloaded successfully.\nSaved as:\n{save_path.name}",
                "Success",
            )
        except Exception as exc:
            logging.error("Spreadsheet download error: %s", exc, exc_info=True)
            pymsgbox.alert(f"Spreadsheet download failed:\n\n{exc}", "Error")

    def create_google_sheet(self, path: Path) -> None:
        try:
            logging.info("Uploading file %s to Google Sheets.", path)
            sheets_workbook = ezsheets.upload(str(path))
            logging.info("Upload completed successfully. ID: %s", sheets_workbook.id)
            pymsgbox.alert(
                text=(
                    "Spreadsheet uploaded successfully to Google Sheets."
                    f"\nTitle: {sheets_workbook.title}"
                ),
                title="Success",
            )
        except Exception as exc:
            logging.error("Google Sheets upload error: %s", exc, exc_info=True)
            pymsgbox.alert(f"Google Sheets upload failed:\n{exc}", "Error")
